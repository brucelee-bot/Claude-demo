from copy import copy
from io import BytesIO
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


TEMPLATE_PATH = Path(__file__).resolve().parent / "templates" / "RD-IP-PS-成果关联详情表.xlsx"


COLUMNS = [
    "rd_code",
    "year",
    "rd_activity",
    "rd_period",
    "ip_code",
    "ip_name",
    "ip_auth_no",
    "ps_display",
    "result_no",
    "result_name",
    "sales_contract_filename",
    "sales_contract_summary",
    "sales_contract_keywords",
    "tech_field_path",
]

COLUMN_TITLES = {
    "rd_code": "RD序号",
    "year": "年份",
    "rd_activity": "研发活动",
    "rd_period": "研发周期",
    "ip_code": "相关知识产权",
    "ip_name": "知识产权名称",
    "ip_auth_no": "授权号",
    "ps_display": "对应PS",
    "result_no": "成果序号",
    "result_name": "成果名称",
    "sales_contract_filename": "销售合同",
    "sales_contract_summary": "销售合同摘要",
    "sales_contract_keywords": "销售合同关键词",
    "tech_field_path": "高新技术领域",
}


def _copy_row_style(ws, source_row, target_row):
    for col in range(1, len(COLUMNS) + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)

    if ws.row_dimensions[source_row].height:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def _clear_template_rows(ws):
    for cell_range in list(ws.merged_cells.ranges):
        if cell_range.min_row >= 2:
            ws.unmerge_cells(str(cell_range))
    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)
    for col in range(1, len(COLUMNS) + 1):
        ws.cell(2, col).value = None


def _merge_year_cells(ws, start_row, rows):
    block_start = None
    current_year = None

    for index, row in enumerate(rows, start=start_row):
        year = row.get("year", "")
        if year != current_year:
            if block_start is not None and index - block_start > 1:
                ws.merge_cells(start_row=block_start, start_column=2, end_row=index - 1, end_column=2)
            block_start = index
            current_year = year

    end_row = start_row + len(rows) - 1
    if block_start is not None and end_row - block_start >= 1:
        ws.merge_cells(start_row=block_start, start_column=2, end_row=end_row, end_column=2)

    for row in range(start_row, end_row + 1):
        ws.cell(row, 2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def export_relation_table(rows, tech_field_path=""):
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active
    for col, key in enumerate(COLUMNS, start=1):
        ws.cell(1, col).value = COLUMN_TITLES.get(key, key)
    _clear_template_rows(ws)

    if rows:
        for offset, row in enumerate(rows):
            excel_row = 2 + offset
            if excel_row != 2:
                ws.insert_rows(excel_row)
            _copy_row_style(ws, 2, excel_row)
            for col, key in enumerate(COLUMNS, start=1):
                value = tech_field_path if key == "tech_field_path" else row.get(key, "")
                ws.cell(excel_row, col).value = value

        _merge_year_cells(ws, 2, rows)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


HEADER_MAP = {
    "序号": "rd_code",
    "rd序号": "rd_code",
    "年份": "year",
    "研发活动": "rd_activity",
    "研发周期": "rd_period",
    "相关知识产权": "ip_code",
    "知识产权": "ip_code",
    "相关知识产权名称": "ip_name",
    "知识产权名称": "ip_name",
    "相关知识产权授权号": "ip_auth_no",
    "授权号": "ip_auth_no",
    "对应ps": "ps_display",
    "对应PS": "ps_display",
    "ps": "ps_display",
    "PS": "ps_display",
    "成果序号": "result_no",
    "成果名称": "result_name",
    "销售合同": "sales_contract_filename",
    "销售合同编号": "sales_contract_code",
    "销售合同文件": "sales_contract_filename",
    "销售合同摘要": "sales_contract_summary",
    "销售合同关键词": "sales_contract_keywords",
    "合同关键词": "sales_contract_keywords",
    "高新技术领域": "tech_field_path",
    "国家重点支持的高新技术领域": "tech_field_path",
}


def _clean_header(value):
    return re.sub(r"\s+", "", str(value or "").strip())


def _cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _merged_value(ws, row, col):
    cell = ws.cell(row, col)
    if cell.value is not None:
        return cell.value
    for cell_range in ws.merged_cells.ranges:
        if cell_range.min_row <= row <= cell_range.max_row and cell_range.min_col <= col <= cell_range.max_col:
            return ws.cell(cell_range.min_row, cell_range.min_col).value
    return None


def _split_ps(value):
    text = _cell_text(value)
    match = re.match(r"^(PS\s*\d+)[\-－—–_\s]+(.+)$", text, re.IGNORECASE)
    if match:
        return match.group(1).replace(" ", "").upper(), match.group(2).strip()
    if re.match(r"^PS\s*\d+$", text, re.IGNORECASE):
        return text.replace(" ", "").upper(), ""
    return "", text


def import_relation_table(file_obj):
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active

    header_row = None
    column_map = {}
    for row in range(1, min(ws.max_row, 10) + 1):
        found = {}
        for col in range(1, ws.max_column + 1):
            header = _clean_header(ws.cell(row, col).value)
            key = HEADER_MAP.get(header)
            if key:
                found[key] = col
        if {"rd_code", "year", "rd_activity", "ip_name", "result_name"}.issubset(found):
            header_row = row
            column_map = found
            break

    if header_row is None:
        raise ValueError("未识别到 RD-IP-PS-成果关联详情表表头")

    rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        raw = {}
        for key, col in column_map.items():
            raw[key] = _cell_text(_merged_value(ws, row_idx, col))
        if not any(raw.values()):
            continue
        ps_code, ps_name = _split_ps(raw.pop("ps_display", ""))
        raw["ps_code"] = ps_code
        raw["ps_name"] = ps_name
        rows.append(raw)

    return rows
