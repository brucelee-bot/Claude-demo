import io
import json
import os
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from flask import Flask, session
from flask_login import AnonymousUserMixin, LoginManager
from openpyxl import Workbook, load_workbook

from modules.docgen.routes import (
    GAOXIN_HR_STAFF_HEADERS,
    _import_hr_staff_excel,
    _normalize_relation_rows,
    _relation_sales_contract_options,
    _summarize_hr_staff_rows,
)
from modules.docgen.relation_table_exporter import export_relation_table, import_relation_table
from modules.docgen.sales_contracts import (
    ensure_sales_contract_codes,
    remap_sales_contract_rows,
    selectable_sales_contracts,
)
from modules.parser.routes import _ip_cert_store, _required_material_store, parser_bp
from modules.scoring.routes import (
    _persist_required_materials,
    _seed_required_material_relation_rows,
)


ROOT = Path(__file__).resolve().parents[1]


def staff_workbook(headers, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


class RequiredMaterialTemplateTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.upload_template = (ROOT / "templates" / "_upload_component.html").read_text(
            encoding="utf-8"
        )
        cls.score_template = (ROOT / "templates" / "score_gaoxin_form.html").read_text(
            encoding="utf-8"
        )
        cls.relation_template = (
            ROOT / "templates" / "application_gaoxin_relation_table.html"
        ).read_text(encoding="utf-8")

    def test_required_material_card_has_four_upload_sections(self):
        self.assertIn("<h3>必须上传的材料</h3>", self.upload_template)
        self.assertEqual(
            self.upload_template.count('class="required-material-section"'),
            4,
        )
        for label in ("财务报表", "专利证书", "人员清单", "销售合同"):
            self.assertIn(f"<strong>{label}</strong>", self.upload_template)

    def test_patent_upload_only_exists_in_required_material_card(self):
        self.assertEqual(self.upload_template.count('id="ip-pdf-file"'), 1)
        self.assertNotIn("上传专利证书 PDF（可多次上传）", self.score_template)

    def test_frontend_calls_staff_and_contract_parser_endpoints(self):
        self.assertIn("fetch('/parser/upload_staff_list'", self.upload_template)
        self.assertIn("fetch('/parser/upload_sales_contract'", self.upload_template)
        self.assertIn("fetch('/parser/sales_contract_upload_ticket'", self.upload_template)
        self.assertIn("fetch('/parser/register_sales_contract'", self.upload_template)
        self.assertIn("https://blob.vercel-storage.com/", self.upload_template)
        self.assertIn("fetch('/parser/sales_contracts'", self.upload_template)
        self.assertIn("/parser/required_materials/finance/", self.upload_template)
        self.assertIn("/parser/required_materials/staff/", self.upload_template)
        self.assertIn("/parser/required_materials/sales_contracts/", self.upload_template)
        self.assertIn("deletePatentCertificate", self.score_template)

    def test_sales_contract_copy_states_that_upload_only_saves_files(self):
        self.assertIn("销售合同仅保存，后续按需解析", self.upload_template)
        self.assertIn("此处不解析合同内容", self.upload_template)
        self.assertIn("正在保存 ${year} 年合同", self.upload_template)
        self.assertNotIn("正在识别 ${year} 年合同", self.upload_template)

    def test_sales_contracts_are_displayed_by_original_pdf_filename(self):
        self.assertIn(
            "materialEscapeHtml(item.original_filename || '销售合同.pdf')",
            self.upload_template,
        )
        self.assertNotIn(
            "item.contract_code || '未编号合同'",
            self.upload_template,
        )
        self.assertIn("已按 PDF 文件名保存", self.upload_template)
        self.assertIn(
            "option.textContent = contract.original_filename || contract.code || '未命名合同';",
            self.relation_template,
        )
        self.assertIn(
            "contract.original_filename === row.sales_contract_filename",
            self.relation_template,
        )
        self.assertNotIn("请选择合同编号", self.relation_template)

    def test_staff_section_exposes_template_download_and_excel_upload(self):
        self.assertIn(
            "url_for('parser.staff_list_template')",
            self.upload_template,
        )
        self.assertIn("下载 Excel 模板", self.upload_template)
        self.assertIn("下载模板填写后，拖拽或点击上传 Excel", self.upload_template)
        self.assertIn("uploadStaffList(this.files)", self.upload_template)

    def test_materials_use_drag_and_click_upload_patterns(self):
        self.assertEqual(
            self.upload_template.count('class="upload-zone material-drop-zone"'),
            3,
        )
        self.assertEqual(
            self.upload_template.count(
                'class="upload-zone material-drop-zone sales-contract-drop-zone"'
            ),
            3,
        )
        self.assertNotIn("material-upload-actions", self.upload_template)
        self.assertIn("height: 152px;", self.upload_template)
        self.assertIn("height: 172px;", self.upload_template)
        for input_id, handler in (
            ("unified-file", "handleFiles"),
            ("ip-pdf-file", "uploadPatentPdf"),
            ("staff-list-file", "uploadStaffList"),
        ):
            self.assertIn(f"id=\"{input_id}\"", self.upload_template)
            self.assertIn(f"onchange=\"{handler}(this.files)\"", self.upload_template)
        for year in ("2023", "2024", "2025"):
            self.assertIn(f"{year}年销售合同", self.upload_template)
            self.assertIn(f'id="sales-contract-files-{year}"', self.upload_template)
            self.assertIn(
                f"onchange=\"uploadSalesContracts('{year}', this.files)\"",
                self.upload_template,
            )
        self.assertNotIn("MAX_SALES_CONTRACTS_PER_YEAR", self.upload_template)
        self.assertIn("按年度保存，可上传多份 PDF", self.upload_template)
        self.assertIn("fd.append('year', year);", self.upload_template)


class RequiredMaterialStaffImportTests(unittest.TestCase):
    def test_staff_template_download_uses_attachment_staff_headers(self):
        app = Flask(__name__)
        app.secret_key = "test-secret"
        app.config["LOGIN_DISABLED"] = True
        login_manager = LoginManager(app)

        @login_manager.user_loader
        def load_user(_user_id):
            return None

        app.register_blueprint(parser_bp, url_prefix="/parser")

        response = app.test_client().get("/parser/staff_list_template")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.mimetype,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(
            "attachment;",
            response.headers.get("Content-Disposition", ""),
        )
        workbook = load_workbook(io.BytesIO(response.data), data_only=True)
        headers = [cell.value for cell in workbook.active[1]]
        self.assertEqual(headers, GAOXIN_HR_STAFF_HEADERS)

    def test_current_staff_format_imports_title_without_removed_fields(self):
        upload = staff_workbook(
            [
                "序号",
                "姓名",
                "身份证号",
                "是否签订合同",
                "是否缴纳社保",
                "学历",
                "职称",
                "是否科技人员",
            ],
            [[1, "张三", "110101199001010011", "是", "是", "本科", "中级工程师", "是"]],
        )

        rows = _import_hr_staff_excel(upload)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["姓名"], "张三")
        self.assertEqual(rows[0]["职称"], "中级工程师")
        self.assertNotIn("入职时间", rows[0])
        self.assertNotIn("工作性质", rows[0])
        summary = _summarize_hr_staff_rows(rows)
        self.assertEqual(summary["hr_total"], 1)
        self.assertEqual(summary["tech_staff"], 1)
        self.assertEqual(summary["title_mid"], 1)

    def test_legacy_staff_format_remains_compatible(self):
        upload = staff_workbook(
            [
                "序号",
                "姓名",
                "身份证号",
                "是否签订合同",
                "入职时间",
                "是否缴纳社保",
                "工作性质",
                "学历",
                "是否科技人员",
            ],
            [[1, "李四", "110101199202020022", "是", "2024-01-01", "是", "研发", "硕士", "是"]],
        )

        rows = _import_hr_staff_excel(upload)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["姓名"], "李四")
        self.assertEqual(rows[0]["学历"], "硕士")
        self.assertEqual(rows[0]["职称"], "")


class RequiredMaterialSalesContractTests(unittest.TestCase):
    def test_large_pdf_uses_direct_blob_upload_and_registers_without_parsing(self):
        class TestAnonymousUser(AnonymousUserMixin):
            id = 1

        app = Flask(__name__)
        app.secret_key = "test-secret"
        app.config["LOGIN_DISABLED"] = True
        login_manager = LoginManager(app)
        login_manager.anonymous_user = TestAnonymousUser

        @login_manager.user_loader
        def load_user(_user_id):
            return None

        app.register_blueprint(parser_bp, url_prefix="/parser")

        with (
            patch("modules.parser.routes.blob_enabled", return_value=True),
            patch(
                "modules.parser.routes.generate_client_upload_token",
                side_effect=lambda relative_path: {
                    "client_token": "vercel_blob_client_store_token",
                    "pathname": f"declare-assistant/{relative_path}",
                    "store_id": "store",
                },
            ),
            patch(
                "modules.parser.routes.blob_metadata",
                return_value={
                    "url": "https://example.private.blob.vercel-storage.com/contract.pdf",
                    "downloadUrl": "https://example.private.blob.vercel-storage.com/contract.pdf?download=1",
                    "etag": "blob-etag",
                    "size": 6 * 1024 * 1024,
                },
            ),
            patch("modules.docgen.routes._extract_pdf_text") as extract_text,
            patch("modules.docgen.routes._extract_sales_contract_info") as extract_info,
            patch.dict(_required_material_store, {}, clear=True),
        ):
            client = app.test_client()
            store_key = "direct-upload-more-than-four"
            with client.session_transaction() as client_session:
                client_session["required_material_store_key"] = store_key
            _required_material_store[store_key] = {
                "finance": [],
                "staff": {},
                "sales_contracts": [
                    {
                        "id": f"existing-{index}",
                        "year": "2024",
                        "contract_sequence": index,
                        "contract_code": f"2024合同{index:02d}",
                        "original_filename": f"已有合同{index}.pdf",
                    }
                    for index in range(1, 5)
                ],
            }
            ticket_response = client.post(
                "/parser/sales_contract_upload_ticket",
                json={
                    "year": "2024",
                    "filename": "无法解析但可以保存的扫描合同.pdf",
                    "size": 6 * 1024 * 1024,
                },
            )
            self.assertEqual(ticket_response.status_code, 200)
            ticket = ticket_response.get_json()
            self.assertTrue(ticket["direct_upload"])
            self.assertEqual(
                ticket["upload"]["original_filename"],
                "无法解析但可以保存的扫描合同.pdf",
            )

            register_response = client.post(
                "/parser/register_sales_contract",
                json={
                    "id": ticket["upload"]["id"],
                    "year": "2024",
                    "filename": "无法解析但可以保存的扫描合同.pdf",
                    "relative_path": ticket["upload"]["relative_path"],
                    "size": 6 * 1024 * 1024,
                },
            )

            self.assertEqual(register_response.status_code, 200)
            saved = register_response.get_json()["file"]
            self.assertEqual(saved["contract_code"], "2024合同05")
            self.assertEqual(saved["summary"], "")
            self.assertEqual(saved["keywords"], "")
            self.assertEqual(saved["size"], 6 * 1024 * 1024)
            self.assertEqual(register_response.get_json()["count"], 5)
            extract_text.assert_not_called()
            extract_info.assert_not_called()

    def test_each_contract_year_accepts_more_than_four_pdfs(self):
        class TestAnonymousUser(AnonymousUserMixin):
            id = 1

        app = Flask(__name__)
        app.secret_key = "test-secret"
        app.config["LOGIN_DISABLED"] = True
        login_manager = LoginManager(app)
        login_manager.anonymous_user = TestAnonymousUser

        @login_manager.user_loader
        def load_user(_user_id):
            return None

        app.register_blueprint(parser_bp, url_prefix="/parser")

        with (
            tempfile.TemporaryDirectory() as upload_root,
            patch("modules.parser.routes.Config.UPLOAD_FOLDER", upload_root),
            patch("modules.docgen.routes._extract_pdf_text", return_value="合同正文") as extract_text,
            patch(
                "modules.docgen.routes._extract_sales_contract_info",
                return_value={"summary": "合同摘要", "keywords": "技术关键词"},
            ) as extract_info,
            patch.dict(_required_material_store, {}, clear=True),
        ):
            client = app.test_client()
            for index in range(1, 7):
                response = client.post(
                    "/parser/upload_sales_contract",
                    data={
                        "year": "2023",
                        "file": (
                            io.BytesIO(f"%PDF-2023-{index}".encode()),
                            f"2023年销售合同{index}.pdf",
                        ),
                    },
                    content_type="multipart/form-data",
                )
                self.assertEqual(response.status_code, 200)
                payload = response.get_json()
                self.assertEqual(payload["file"]["year"], "2023")
                self.assertEqual(
                    payload["file"]["contract_code"],
                    f"2023合同{index:02d}",
                )
                self.assertEqual(payload["file"]["summary"], "")
                self.assertEqual(payload["file"]["keywords"], "")
                self.assertEqual(payload["count"], index)

            next_year = client.post(
                "/parser/upload_sales_contract",
                data={
                    "year": "2024",
                    "file": (io.BytesIO(b"%PDF-2024-1"), "2024年销售合同1.pdf"),
                },
                content_type="multipart/form-data",
            )
            self.assertEqual(next_year.status_code, 200)
            self.assertEqual(next_year.get_json()["file"]["year"], "2024")
            self.assertEqual(
                next_year.get_json()["file"]["contract_code"],
                "2024合同01",
            )
            self.assertEqual(next_year.get_json()["count"], 1)

            restored = client.get("/parser/sales_contracts")
            self.assertEqual(restored.status_code, 200)
            restored_contracts = restored.get_json()["contracts"]
            self.assertEqual(
                [item["contract_code"] for item in restored_contracts],
                [
                    "2023合同01",
                    "2023合同02",
                    "2023合同03",
                    "2023合同04",
                    "2023合同05",
                    "2023合同06",
                    "2024合同01",
                ],
            )
            extract_text.assert_not_called()
            extract_info.assert_not_called()

    def test_saved_contract_can_be_deleted_with_its_file(self):
        class TestAnonymousUser(AnonymousUserMixin):
            id = 1

        app = Flask(__name__)
        app.secret_key = "test-secret"
        app.config["LOGIN_DISABLED"] = True
        login_manager = LoginManager(app)
        login_manager.anonymous_user = TestAnonymousUser

        @login_manager.user_loader
        def load_user(_user_id):
            return None

        app.register_blueprint(parser_bp, url_prefix="/parser")

        with (
            tempfile.TemporaryDirectory() as upload_root,
            patch("modules.parser.routes.Config.UPLOAD_FOLDER", upload_root),
            patch.dict(_required_material_store, {}, clear=True),
        ):
            client = app.test_client()
            uploaded = client.post(
                "/parser/upload_sales_contract",
                data={
                    "year": "2024",
                    "file": (io.BytesIO(b"%PDF-1.4 saved contract"), "合同.pdf"),
                },
                content_type="multipart/form-data",
            )
            self.assertEqual(uploaded.status_code, 200)
            file_meta = uploaded.get_json()["file"]
            stored_path = os.path.join(upload_root, file_meta["relative_path"])
            self.assertTrue(os.path.exists(stored_path))

            deleted = client.delete(
                f"/parser/required_materials/sales_contracts/{file_meta['id']}"
            )

            self.assertEqual(deleted.status_code, 200)
            self.assertEqual(deleted.get_json()["remaining"], 0)
            self.assertFalse(os.path.exists(stored_path))
            restored = client.get("/parser/sales_contracts").get_json()["contracts"]
            self.assertEqual(restored, [])


class RequiredMaterialDeletionTests(unittest.TestCase):
    def setUp(self):
        class TestAnonymousUser(AnonymousUserMixin):
            id = 1

        self.app = Flask(__name__)
        self.app.secret_key = "test-secret"
        self.app.config["LOGIN_DISABLED"] = True
        login_manager = LoginManager(self.app)
        login_manager.anonymous_user = TestAnonymousUser

        @login_manager.user_loader
        def load_user(_user_id):
            return None

        self.app.register_blueprint(parser_bp, url_prefix="/parser")

    def test_finance_and_staff_materials_delete_by_id(self):
        store_key = "materials-delete-test"
        materials = {
            "finance": [
                {
                    "id": "finance-1",
                    "filename": "2024.xlsx",
                    "data": {"fin_2024_revenue": "100"},
                    "validation": {"verified": True},
                },
                {
                    "id": "finance-2",
                    "filename": "2025.xlsx",
                    "data": {"fin_2025_revenue": "200"},
                    "validation": {"verified": True},
                },
            ],
            "staff": {
                "id": "staff-1",
                "filename": "人员清单.xlsx",
                "rows": [{"姓名": "张三"}],
                "summary": {"hr_total": 1},
            },
            "sales_contracts": [],
        }

        with patch.dict(_required_material_store, {store_key: materials}, clear=True):
            client = self.app.test_client()
            with client.session_transaction() as client_session:
                client_session["required_material_store_key"] = store_key
                client_session["last_finance_data"] = {
                    "fin_2024_revenue": "100",
                    "fin_2025_revenue": "200",
                }

            finance_deleted = client.delete(
                "/parser/required_materials/finance/finance-2"
            )
            self.assertEqual(finance_deleted.status_code, 200)
            self.assertEqual(
                finance_deleted.get_json()["data"],
                {"fin_2024_revenue": "100"},
            )
            with client.session_transaction() as client_session:
                self.assertEqual(
                    client_session["last_finance_data"],
                    {"fin_2024_revenue": "100"},
                )

            staff_deleted = client.delete(
                "/parser/required_materials/staff/staff-1"
            )
            self.assertEqual(staff_deleted.status_code, 200)
            self.assertEqual(materials["staff"], {})

    def test_patent_delete_uses_stable_id_and_removes_staged_pdf(self):
        store_key = "ip-delete-test"
        with (
            tempfile.TemporaryDirectory() as upload_root,
            patch("modules.parser.routes.Config.UPLOAD_FOLDER", upload_root),
            patch.dict(_ip_cert_store, {}, clear=True),
        ):
            relative_path = os.path.join("ip_pending", "patent.pdf")
            staged_path = os.path.join(upload_root, relative_path)
            os.makedirs(os.path.dirname(staged_path), exist_ok=True)
            Path(staged_path).write_bytes(b"%PDF-1.4 patent")
            _ip_cert_store[store_key] = [
                {
                    "id": "patent-1",
                    "filename": "专利证书.pdf",
                    "parsed": {
                        "patent_type": "invention",
                        "details": {"name": "测试专利"},
                    },
                    "source_pdf": {
                        "relative_path": relative_path,
                        "sync_status": "staged",
                    },
                }
            ]

            client = self.app.test_client()
            with client.session_transaction() as client_session:
                client_session["ip_cert_store_key"] = store_key

            deleted = client.post(
                "/parser/delete_ip",
                json={"id": "patent-1"},
            )

            self.assertEqual(deleted.status_code, 200)
            self.assertEqual(deleted.get_json()["remaining"], 0)
            self.assertEqual(deleted.get_json()["certificates"], [])
            self.assertFalse(os.path.exists(staged_path))

    def test_legacy_contracts_receive_stable_yearly_codes(self):
        contracts = [
            {"id": "a", "year": "2023"},
            {"id": "b", "year": "2024"},
            {"id": "c", "year": "2023"},
        ]

        ensure_sales_contract_codes(contracts)
        self.assertEqual(
            [item["contract_code"] for item in contracts],
            ["2023合同01", "2024合同01", "2023合同02"],
        )
        ensure_sales_contract_codes(contracts)
        self.assertEqual(
            [item["contract_code"] for item in contracts],
            ["2023合同01", "2024合同01", "2023合同02"],
        )

    def test_legacy_contract_years_are_inferred_and_duplicate_hashes_do_not_use_slots(self):
        contracts = [
            {"id": "a", "original_filename": "2023-25合同.pdf", "sha256": "hash-a"},
            {"id": "b", "original_filename": "2023-31合同.pdf", "sha256": "hash-b"},
            {"id": "b-copy", "original_filename": "副本.pdf", "sha256": "hash-b"},
            {"id": "c", "stored_filename": "uuid_2023-54合同.pdf", "sha256": "hash-c"},
            {"id": "d", "relative_path": "contracts/2023-100合同.pdf", "sha256": "hash-d"},
            {"id": "e", "original_filename": "第五份.pdf", "sha256": "hash-e"},
        ]
        relation_rows = [
            {"year": "2023", "sales_contract_file_id": "e"},
            {"year": "2023", "sales_contract_file_id": "b-copy"},
        ]

        ensure_sales_contract_codes(contracts, relation_rows)
        options = selectable_sales_contracts(contracts, relation_rows)
        relation_options = _relation_sales_contract_options(contracts)
        remap_sales_contract_rows(relation_rows, contracts)

        expected_codes = [
            "2023合同01",
            "2023合同02",
            "2023合同03",
            "2023合同04",
            "2023合同05",
        ]
        self.assertEqual([item["contract_code"] for item in options], expected_codes)
        self.assertEqual([item["code"] for item in relation_options], expected_codes)
        self.assertEqual(contracts[2]["duplicate_of"], "b")
        self.assertEqual(contracts[2]["contract_code"], "2023合同02")
        self.assertEqual(contracts[5]["contract_code"], "2023合同05")
        self.assertEqual(relation_rows[1]["sales_contract_file_id"], "b")
        self.assertEqual(relation_rows[1]["sales_contract_code"], "2023合同02")

    def test_filename_year_takes_priority_over_relation_row_year(self):
        contracts = [
            {
                "id": "contract-2024",
                "original_filename": "2024-06销售合同.pdf",
            }
        ]
        relation_rows = [
            {
                "year": "2023",
                "sales_contract_file_id": "contract-2024",
            }
        ]

        ensure_sales_contract_codes(contracts, relation_rows)

        self.assertEqual(contracts[0]["year"], "2024")
        self.assertEqual(contracts[0]["contract_code"], "2024合同01")


class SalesContractRelationExportTests(unittest.TestCase):
    def test_export_uses_original_pdf_filename_in_sales_contract_column(self):
        stream = export_relation_table(
            [
                {
                    "rd_code": "RD01",
                    "year": "2024",
                    "rd_activity": "智能校核研发",
                    "sales_contract_code": "2024合同01",
                    "sales_contract_filename": "继电保护智能校核服务合同.pdf",
                }
            ]
        )
        workbook = load_workbook(stream, data_only=True)
        sheet = workbook.active

        self.assertEqual(sheet.cell(1, 11).value, "销售合同")
        self.assertEqual(sheet.cell(2, 11).value, "继电保护智能校核服务合同.pdf")

    def test_import_reads_sales_contract_column_as_pdf_filename(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "序号",
                "年份",
                "研发活动",
                "知识产权名称",
                "成果名称",
                "销售合同",
            ]
        )
        sheet.append(
            [
                "RD01",
                "2024",
                "智能校核研发",
                "智能校核专利",
                "智能校核成果",
                "继电保护智能校核服务合同.pdf",
            ]
        )
        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)

        rows = import_relation_table(stream)

        self.assertEqual(
            rows[0]["sales_contract_filename"],
            "继电保护智能校核服务合同.pdf",
        )


class RequiredMaterialPersistenceTests(unittest.TestCase):
    def test_relation_rows_keep_contracts_unselected_until_user_chooses(self):
        company = SimpleNamespace(
            ip_certs_json=json.dumps(
                [
                    {
                        "parsed": {
                            "details": {
                                "name": "一种继电保护校核方法",
                                "patent_no": "ZL202510000001.0",
                            }
                        }
                    }
                ],
                ensure_ascii=False,
            )
        )
        data = {}
        contracts = [
            {
                "id": "contract-1",
                "original_filename": "销售合同.pdf",
                "year": "2023",
                "summary": "销售继电保护校核系统。",
                "keywords": "继电保护；智能校核",
            }
        ]

        _seed_required_material_relation_rows(data, company, contracts)

        rows = data["gaoxin_relation_table"]["rows"]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["ip_code"], "IP01")
        self.assertEqual(rows[0]["ip_name"], "一种继电保护校核方法")
        self.assertEqual(rows[0]["ip_auth_no"], "ZL202510000001.0")
        self.assertEqual(rows[0]["sales_contract_file_id"], "")
        self.assertEqual(rows[0]["sales_contract_keywords"], "")
        self.assertEqual(rows[0]["year"], "2025")
        self.assertEqual(rows[0]["rd_code"], "")
        self.assertEqual(rows[0]["ps_code"], "")
        self.assertEqual(rows[0]["result_name"], "")

    def test_persist_required_materials_writes_staff_contract_and_summary(self):
        app = Flask(__name__)
        app.secret_key = "test-secret"
        with tempfile.TemporaryDirectory() as upload_root:
            app.config["UPLOAD_FOLDER"] = upload_root
            staged_relative_path = os.path.join(
                "required_materials_pending",
                "sales_contracts",
                "contract-1.pdf",
            )
            staged_path = os.path.join(upload_root, staged_relative_path)
            os.makedirs(os.path.dirname(staged_path), exist_ok=True)
            Path(staged_path).write_bytes(b"%PDF-1.4 test contract")

            staff_rows = [
                {
                    "序号": "1",
                    "姓名": "张三",
                    "身份证号": "110101199001010011",
                    "是否签订合同": "是",
                    "是否缴纳社保": "是",
                    "学历": "本科",
                    "职称": "中级工程师",
                    "是否科技人员": "是",
                }
            ]
            materials = {
                "staff": {
                    "filename": "人员清单.xlsx",
                    "rows": staff_rows,
                    "summary": _summarize_hr_staff_rows(staff_rows),
                },
                "sales_contracts": [
                    {
                        "id": "contract-1",
                        "original_filename": "销售合同.pdf",
                        "relative_path": staged_relative_path,
                        "year": "2024",
                        "contract_sequence": 1,
                        "contract_code": "2024合同01",
                        "summary": "销售继电保护校核系统。",
                        "keywords": "继电保护；智能校核",
                        "sha256": "contract-sha256",
                    }
                ],
            }
            company = SimpleNamespace(
                id=13,
                user_id=7,
                data_json="{}",
                ip_certs_json=json.dumps(
                    [
                        {
                            "parsed": {
                                "details": {
                                    "name": "一种继电保护校核方法",
                                    "patent_no": "ZL202510000001.0",
                                }
                            }
                        }
                    ],
                    ensure_ascii=False,
                ),
            )

            with app.test_request_context("/score/gaoxin", method="POST"):
                session["last_finance_data"] = {
                    "fin_2025_revenue": "1000",
                    "fin_2025_net_assets": "500",
                }
                with (
                    patch(
                        "modules.parser.routes._get_required_materials",
                        return_value=materials,
                    ),
                    patch("modules.parser.routes._clear_required_materials") as clear_materials,
                ):
                    _persist_required_materials(company)

            data = json.loads(company.data_json)
            self.assertEqual(data["hr_staff_rows"], staff_rows)
            self.assertEqual(data["staff_total"], 1)
            self.assertEqual(data["tech_staff"], 1)
            self.assertTrue(data["required_materials"]["finance"]["recognized"])
            self.assertEqual(data["required_materials"]["patents"]["count"], 1)
            self.assertEqual(data["required_materials"]["staff"]["count"], 1)
            self.assertEqual(len(data["required_materials"]["sales_contracts"]), 1)
            self.assertEqual(
                data["required_materials"]["sales_contracts"][0]["year"],
                "2024",
            )
            self.assertEqual(
                data["required_materials"]["sales_contracts"][0]["contract_code"],
                "2024合同01",
            )

            files = data["gaoxin_attachments"]["relation_sales_contract"]["files"]
            self.assertEqual(len(files), 1)
            persisted_path = os.path.join(upload_root, files[0]["relative_path"])
            self.assertTrue(os.path.exists(persisted_path))
            self.assertEqual(files[0]["year"], "2024")
            self.assertEqual(files[0]["contract_code"], "2024合同01")

            relation_row = data["gaoxin_relation_table"]["rows"][0]
            self.assertEqual(relation_row["ip_name"], "一种继电保护校核方法")
            self.assertEqual(relation_row["sales_contract_file_id"], "")
            self.assertEqual(relation_row["sales_contract_keywords"], "")
            self.assertEqual(relation_row["year"], "2025")
            clear_materials.assert_called_once_with()

    def test_blob_only_contract_remains_selectable_after_scoring_persistence(self):
        app = Flask(__name__)
        app.secret_key = "test-secret"
        with tempfile.TemporaryDirectory() as upload_root:
            app.config["UPLOAD_FOLDER"] = upload_root
            staged_relative_path = (
                "required_materials_pending/7/materials-blob/"
                "sales_contracts/2024/contract-blob.pdf"
            )
            materials = {
                "staff": {},
                "sales_contracts": [
                    {
                        "id": "contract-blob",
                        "original_filename": "2024年继电保护服务合同.pdf",
                        "stored_filename": "contract-blob.pdf",
                        "relative_path": staged_relative_path,
                        "year": "2024",
                        "contract_sequence": 1,
                        "contract_code": "2024合同01",
                        "summary": "",
                        "keywords": "",
                        "sha256": "",
                        "blob_url": "https://example.private.blob.vercel-storage.com/contract.pdf",
                        "blob_download_url": (
                            "https://example.private.blob.vercel-storage.com/"
                            "contract.pdf?download=1"
                        ),
                        "blob_etag": "blob-etag",
                        "size": 8 * 1024 * 1024,
                    }
                ],
            }
            company = SimpleNamespace(
                id=13,
                user_id=7,
                data_json="{}",
                ip_certs_json="[]",
            )

            with app.test_request_context("/score/gaoxin", method="POST"):
                session["required_material_store_key"] = "materials-blob"
                with (
                    patch.dict(
                        _required_material_store,
                        {"materials-blob": materials},
                        clear=True,
                    ),
                    patch(
                        "modules.docgen.routes._source_upload_path",
                        return_value=os.path.join(upload_root, "missing-contract.pdf"),
                    ),
                    patch("modules.parser.routes.delete_file") as delete_blob,
                ):
                    _persist_required_materials(company)
                    self.assertNotIn("materials-blob", _required_material_store)
                    delete_blob.assert_not_called()

            data = json.loads(company.data_json)
            files = data["gaoxin_attachments"]["relation_sales_contract"]["files"]
            self.assertEqual(len(files), 1)
            self.assertEqual(files[0]["id"], "contract-blob")
            self.assertEqual(files[0]["year"], "2024")
            self.assertEqual(files[0]["contract_code"], "2024合同01")
            self.assertEqual(files[0]["original_filename"], "2024年继电保护服务合同.pdf")
            self.assertEqual(files[0]["relative_path"], staged_relative_path)
            self.assertEqual(files[0]["blob_etag"], "blob-etag")
            self.assertEqual(
                _relation_sales_contract_options(files),
                [
                    {
                        "id": "contract-blob",
                        "code": "2024合同01",
                        "year": "2024",
                        "original_filename": "2024年继电保护服务合同.pdf",
                        "summary": "",
                        "keywords": "",
                    }
                ],
            )
            self.assertEqual(
                data["required_materials"]["sales_contracts"][0]["id"],
                "contract-blob",
            )

    def test_relation_rows_preserve_selected_contract_code(self):
        rows = _normalize_relation_rows(
            [
                {
                    "year": "2024",
                    "rd_code": "RD01",
                    "ip_name": "继电保护专利",
                    "ps_code": "PS01",
                    "result_no": "成果01",
                    "result_name": "继电保护成果",
                    "sales_contract_file_id": "contract-1",
                    "sales_contract_code": "2024合同01",
                    "sales_contract_filename": "销售合同.pdf",
                    "sales_contract_summary": "合同摘要",
                    "sales_contract_keywords": "继电保护；智能校核",
                }
            ]
        )

        self.assertEqual(rows[0]["sales_contract_file_id"], "contract-1")
        self.assertEqual(rows[0]["sales_contract_code"], "2024合同01")


if __name__ == "__main__":
    unittest.main()
